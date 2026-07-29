#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_workbook.py — turn extract.mjs output into a reviewer's .xlsx (plus a text mirror).

    node docs/language-review/pipeline/extract.mjs --lane es-latam --scope taught
    python docs/language-review/pipeline/build_workbook.py --lane es-latam --scope taught

Writes into docs/language-review/<lane>/template/:
    <lane>-<scope>-review-v<version>.xlsx    the file the reviewer receives
    <lane>-<scope>-review-v<version>.md      a text mirror of every row

THREE SCOPES, BECAUSE THEY ARE THREE DIFFERENT JOBS
  interface    does the app read naturally in this language?
  taught       is the language it TEACHES correct — and is every distractor genuinely wrong?
  explanation  is this language's EXPLANATION of another language accurate and natural?
A reviewer brings different attention to each, and the volumes differ by an order of
magnitude. One workbook holding all three would be unreviewable.

THE ROW IS A QUESTION, NOT A STRING
A single question carries a prompt, four options, an explanation and up to four notes.
Split into strings that is eleven rows and one decision; kept whole it is one row and the
same decision, and the reviewer can actually see whether a distractor is secretly correct.

WHY THE .md MIRROR EXISTS
The .xlsx is a zip of XML: git can store it but not diff it, and reading one row means
opening Excel. The mirror is the same content as plain text, so a pull request shows what
changed between generations and `grep` finds a string. Generated, never hand-edited.

ALL REVIEWER-FACING TEXT LIVES IN i18n/<lane>.json — a reviewer should never work through
instructions in a language they are being hired for their fluency in. Adding a lane means
writing that file; the script refuses to fall back to English silently.
"""
import argparse, json, re, sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent.parent

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="3B2A5A")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
IN_FILL = PatternFill("solid", fgColor="FFF7CC")      # reviewer writes here
REF_FILL = PatternFill("solid", fgColor="F4F1F8")     # read-only reference
ALT_FILL = PatternFill("solid", fgColor="FFF0F5")
OUT_FILL = PatternFill("solid", fgColor="EDEDED")     # advisory / out-of-lane
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
BODY_I = Font(name=FONT, size=10, italic=True, color="666666")
_T = Side(style="thin", color="D5CCE4")
BORDER = Border(left=_T, right=_T, top=_T, bottom=_T)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


# Thousands separator is language-dependent, and getting it wrong is not cosmetic:
# "13.554" reads as a decimal in French, where the separator is a narrow no-break space.
# es/pt/de/it use the point; the rest of the lanes we ship get U+202F.
def _thousands(n, lane):
    sep = "." if str(lane).split("-")[0] in ("es", "pt", "de", "it") else "\u202f"
    return f"{n:,}".replace(",", sep)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--lane", default="es-latam")
    ap.add_argument("--scope", default="interface", choices=["interface", "taught", "explanation"])
    ap.add_argument("--data", default=None)
    ap.add_argument("--out-dir", default=None)
    args = ap.parse_args()
    lane, scope = args.lane, args.scope

    data_path = Path(args.data) if args.data else HERE / ".cache" / f"{lane}-{scope}-review-data.json"
    if not data_path.exists():
        sys.exit(f"no extract data at {data_path}\n"
                 f"run: node docs/language-review/pipeline/extract.mjs --lane {lane} --scope {scope}")
    D = json.loads(data_path.read_text(encoding="utf-8"))

    i18n_path = HERE / "i18n" / f"{lane}.json"
    if not i18n_path.exists():
        sys.exit(f"no reviewer-facing copy at {i18n_path}.\n"
                 f"Write it before generating a {lane} packet — see i18n/es-latam.json for the shape. "
                 f"Falling back to another language would hand a native reviewer instructions they "
                 f"cannot read, so this is a hard stop.")
    T = json.loads(i18n_path.read_text(encoding="utf-8"))
    if scope not in T.get("scopes", {}):
        sys.exit(f"{i18n_path} has no copy for scope \"{scope}\" — add it under \"scopes\".")
    SC = T["scopes"][scope]

    version = D.get("currentVersion") or _repo_version()
    out_dir = Path(args.out_dir) if args.out_dir else REPO / "docs" / "language-review" / lane / "template"
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{lane}-{scope}-review-v{version}"

    S, HH = dict(T["sheets"]), T["headers"]
    S.update(SC.get("sheetNames", {}))   # scope-specific tab numbering
    VERD = '"' + ",".join(T["verdicts"]) + '"'
    OKW, CHW, DQW = T["verdicts"]
    wb = Workbook(); wb.remove(wb.active)
    mirror, built = [], []          # built: (sheet_key, sheet_name, rows, verdict_col)

    corpus_size = D.get("corpusSize", 0)

    def _sheet_token(m):
        key = m.group(1)
        if key not in S:
            sys.exit(f'{i18n_path} scope "{scope}" uses {{sheet:{key}}}, but "{key}" is not a '
                     f'sheet key. Known keys: {", ".join(sorted(S))}')
        return S[key]

    def sub(text):
        # A missing corpusSize used to render silently as "0", so reviewer copy could ship
        # reading "a sample of 40 per language, of about 0 in total". Copy that cites the
        # corpus has to be built against a scope that actually counts one.
        if "{corpus}" in text and not corpus_size:
            sys.exit(
                f'{i18n_path} scope "{scope}" uses {{corpus}}, but this extract carries no '
                f'corpusSize - it would render as "0" in the reviewer\'s instructions. '
                f'Either drop {{corpus}} from that string, or make extract.mjs emit '
                f'corpusSize for the {scope} scope.')
        # {sheet:KEY} resolves to whatever this scope numbered that tab. Copy that names a tab
        # by hand goes stale the moment a scope renumbers (which is how the es-latam interface
        # packet came to cite an "8-Muestra-contenido" tab it does not contain). Whether the
        # tab is actually built is checked once, below, when every sheet exists.
        text = re.sub(r"\{sheet:([A-Za-z]+)\}", _sheet_token, text)
        return (text.replace("{lane}", D["laneLabel"]).replace("{version}", version)
                    .replace("{scope}", SC["title"])
                    .replace("{corpus}", _thousands(corpus_size, lane)))

    def finish(ws, ncols, widths, nrows, dv_col, dv_formula=None):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
        # A section can legitimately have ZERO rows: the fr-fr lane has no
        # data/vocab/<x>Words.fr.js glossaries, because the French Word Bank glosses were
        # translated in place inside data/tracks/l10n/*.fr.js instead of into separate
        # files. openpyxl builds the validation range as "<col>2:<col><nrows+1>", which is
        # "D2:D1" when nrows is 0 and raises ValueError. Header-only sheet, no dropdown.
        if nrows < 1:
            return
        dv = DataValidation(type="list", formula1=dv_formula or VERD, allow_blank=True, showDropDown=False)
        dv.error = dv.prompt = T["validationMsg"]
        dv.errorTitle = dv.promptTitle = T["validationTitle"]
        ws.add_data_validation(dv); dv.add(f"{dv_col}2:{dv_col}{nrows + 1}")

    def paint(ws, ncols, input_cols, wrap_cols, out_of_scope=None):
        for ri in range(2, ws.max_row + 1):
            oos = bool(out_of_scope and out_of_scope[ri - 2])
            for ci in range(1, ncols + 1):
                c = ws.cell(row=ri, column=ci)
                c.font = BODY_I if oos else BODY
                c.border = BORDER
                c.alignment = WRAP if ci in wrap_cols else TOP
                c.fill = IN_FILL if ci in input_cols else (
                    OUT_FILL if oos else (ALT_FILL if ri % 2 == 0 else REF_FILL))

    def emit(key, headers, rows, widths, input_cols, wrap_cols, dv_col,
             dv_formula=None, out_of_scope=None):
        name = S[key]
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        paint(ws, len(headers), input_cols, wrap_cols, out_of_scope)
        finish(ws, len(headers), widths, len(rows), dv_col, dv_formula)
        mirror.append((name, headers, rows))
        built.append((key, name, len(rows), dv_col))
        return ws

    def decisions_sheet(key):
        HD = HH["decisions"]
        body = [[f"DEC-{i:02d}", sub(d[0]), sub(d[1]), sub(d[2]), None, None, None]
                for i, d in enumerate(SC["decisions"], start=1)]
        ws = emit(key, HD, body, [11, 30, 76, 62, 16, 46, 42], {5, 6, 7}, {2, 3, 4, 6, 7},
                  "E", '"' + ",".join(T["decisionVerdicts"]) + '"')
        for ri in range(2, ws.max_row + 1):
            ws.row_dimensions[ri].height = 62

    sect = D["sections"]

    # ================================================================ interface sheets
    if scope == "interface":
        HT = HH["text"]
        TW = [11, 26, 32, 58, 58, 14, 46, 40, 11, 46, 30]

        def text_sheet(key, rows, pri):
            body = []
            for r in rows:
                n = T["notes"].get(str(r.get("key", "")), "")
                body.append([r["id"], T["groups"].get(r.get("group", ""), r.get("group", "")),
                             r.get("key", ""), r.get("en", ""), r.get("tgt", ""), None, None, None,
                             T["priority"]["high"] if n else T["priority"][pri], n, r.get("file", "")])
            emit(key, HT, body, TW, {6, 7, 8}, {4, 5, 7, 8, 10, 11}, "F")

        HV = HH["variants"]
        vb, oos = [], []
        for r in sect["variants"]:
            scope_lbl = r["scope"] + ("" if r["inScope"] else f'  ({T["advisory"]})')
            vb.append([r["id"], r["n"], r["gloss"], T["variantKinds"][r["kind"]], r["term"],
                       r["countries"], r["label"], scope_lbl, None, None, None, None])
            oos.append(not r["inScope"])
        emit("variants", HV, vb, [13, 11, 26, 22, 26, 38, 20, 24, 14, 26, 32, 44],
             {9, 10, 11, 12}, {5, 6, 7, 10, 11, 12}, "I", out_of_scope=oos)
        decisions_sheet("decisions")
        text_sheet("variantConfig", sect["variantConfig"], "high")
        text_sheet("security", sect["security"], "high")
        text_sheet("ui", sect["ui"], "medium")
        text_sheet("components", sect["components"], "medium")
        text_sheet("prose", sect["prose"], "medium")
        text_sheet("changelog", sect["changelog"], "low")

    # ================================================================ content sheets
    else:
        HQ = HH["questions"]
        emit("questions", HQ,
             [[q["id"], q["track"], q["cat"], q["difficulty"], q["key"], q["prompt"], q["correct"],
               q["distractors"], q["explain"], q["explainOther"], q["notes"], None, None, None, q["file"]]
              for q in sect.get("questions", [])],
             [11, 15, 12, 8, 13, 46, 26, 40, 52, 46, 44, 14, 46, 40, 26],
             {12, 13, 14}, {6, 7, 8, 9, 10, 11, 13, 14}, "L")

        HF = HH["fono"]
        emit("fono", HF,
             [[f["id"], f["track"], f["key"], f["difficulty"], f["text"], f["sound"], f["correct"],
               f["distractors"], f["explain"], f["pairs"], None, None, None, f["file"]]
              for f in sect.get("fono", [])],
             [11, 15, 15, 8, 34, 30, 30, 42, 50, 46, 14, 46, 40, 26],
             {11, 12, 13}, {5, 6, 7, 8, 9, 10, 12, 13}, "K")

        vocab = sect.get("vocab", [])
        if scope == "taught":
            HB = HH["vocabTaught"]
            emit("vocab", HB,
                 [[v["id"], v["key"], v["word"], v["gloss"], v["pos"], v["tier"], v["note"],
                   None, None, None, v["file"]]
                  for v in vocab],
                 [11, 10, 30, 36, 14, 8, 52, 14, 34, 40, 28],
                 {8, 9, 10}, {3, 4, 7, 9, 10}, "H")
        else:
            HB = HH["vocabGlossary"]
            emit("vocab", HB,
                 [[v["id"], v["key"], v["source"], v["sourceGloss"], v["gloss"], v["tier"],
                   None, None, None, v["file"]]
                  for v in vocab],
                 [11, 10, 30, 36, 32, 8, 14, 34, 40, 28],
                 {7, 8, 9}, {3, 4, 5, 8, 9}, "G")

        if sect.get("overlaySample"):
            HO = HH["overlay"]
            emit("overlay", HO,
                 [[o["id"], o["track"], o["key"], o["prompt"], o["options"], o["notes"],
                   None, None, None, o["file"]] for o in sect["overlaySample"]],
                 [11, 20, 14, 50, 46, 50, 14, 46, 40, 32],
                 {7, 8, 9}, {4, 5, 6, 8, 9, 10}, "G")
        decisions_sheet("decisionsContent")

    # ======================================= stale tab-reference lint
    # Scopes renumber their tabs (interface drops the sample sheet and moves the changelog
    # from 9 to 8), but reviewer copy is written once and drifts. Any tab named in the copy
    # that this scope does not build is an instruction the reviewer cannot act on, so fail
    # the build rather than ship it. Runs here because it needs the full built set.
    _built_names = [name for _, name, _, _ in built] + [S["readme"], S["summary"]]
    _absent = {n for n in S.values() if n not in _built_names}
    _stale = []
    for section in ("readme", "decisions", "example"):
        for idx, entry in enumerate(SC.get(section, [])):
            for cell in (entry if isinstance(entry, (list, tuple)) else [entry]):
                if not isinstance(cell, str):
                    continue
                for key in re.findall(r"\{sheet:([A-Za-z]+)\}", cell):
                    if key in S and S[key] not in _built_names:
                        _stale.append(f'  {section}[{idx}]: {{sheet:{key}}} -> "{S[key]}"')
                for name in _absent:
                    if name in cell:
                        _stale.append(f'  {section}[{idx}]: literal "{name}"')
    if _stale:
        sys.exit(f'{i18n_path} scope "{scope}" points the reviewer at tabs this packet does '
                 f'not build:\n' + "\n".join(_stale)
                 + f'\n\nBuilt: {", ".join(_built_names)}\n'
                   f'Fix the copy, or use {{sheet:KEY}} so the name follows the scope.')

    # ================================================================ LÉEME
    ws = wb.create_sheet(S["readme"])
    ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 118
    r = 1
    for kind, text in SC["readme"]:
        text = sub(text)
        c = ws.cell(row=r, column=2,
                    value=("• " + text) if kind == "b" else ("    – " + text) if kind == "i" else text)
        c.alignment = WRAP
        if kind == "h1":
            c.font = Font(name=FONT, size=16, bold=True, color="3B2A5A"); ws.row_dimensions[r].height = 26
        elif kind == "h2":
            c.value = None; r += 1
            c = ws.cell(row=r, column=2, value=text)
            c.font = Font(name=FONT, size=12, bold=True, color="7A4E9B")
            c.alignment = WRAP; ws.row_dimensions[r].height = 22
        else:
            c.font = BODY
            ws.row_dimensions[r].height = max(15, 13 * (len(text) // 105 + 1))
        r += 1
        if kind == "h2" and text == T["exampleAnchor"]:
            ex = SC["example"]
            for i, h in enumerate(ex["headers"]):
                cc = ws.cell(row=r, column=2 + i, value=h)
                cc.font = HDR_FONT; cc.fill = HDR_FILL; cc.border = BORDER
                cc.alignment = Alignment(vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(2 + i)].width = 118 if i == 0 else 34
            for i, v in enumerate(ex["row"]):
                cc = ws.cell(row=r + 1, column=2 + i, value=v)
                cc.font = BODY; cc.border = BORDER; cc.alignment = WRAP
                cc.fill = IN_FILL if i >= 1 else REF_FILL
            ws.row_dimensions[r].height = 28; ws.row_dimensions[r + 1].height = 52
            r += 3
            ws.column_dimensions["B"].width = 118
    ws.sheet_view.showGridLines = False

    # ================================================================ RESUMEN
    ws = wb.create_sheet(S["summary"])
    ws.column_dimensions["A"].width = 8
    for col, w in zip("BCDEFGH", [30, 62, 10, 10, 12, 10, 13]):
        ws.column_dimensions[col].width = w
    ws["B2"] = sub(T["summaryTitle"])
    ws["B2"].font = Font(name=FONT, size=16, bold=True, color="3B2A5A")
    ws["B3"] = SC["subtitle"]; ws["B3"].font = Font(name=FONT, size=11, italic=True, color="666666")
    for i, lbl in enumerate(T["signoffFields"]):
        rr = 5 + i
        ws.cell(row=rr, column=2, value=lbl).font = BODY_B
        c = ws.cell(row=rr, column=3); c.fill = IN_FILL; c.border = BORDER; c.font = BODY
    dv = DataValidation(type="list", formula1='"' + ",".join(T["signoffValues"]) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv); dv.add(f"C{5 + len(T['signoffFields']) - 1}")

    start = 11
    for i, h in enumerate(HH["summary"]):
        c = ws.cell(row=start, column=1 + i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BORDER
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[start].height = 30
    for i, (key, name, n, vcol) in enumerate(built):
        rr = start + 1 + i
        q = f"'{name}'!${vcol}$2:${vcol}${n + 1}"
        ok = T["decisionVerdicts"][0] if key.startswith("decisions") else OKW
        ws.cell(row=rr, column=1, value=i + 1)
        ws.cell(row=rr, column=2, value=name)
        ws.cell(row=rr, column=3, value=sub(T["sheetBlurbs"].get(key, T["sheetBlurbs"].get("decisions", "")) if not key.startswith("decisions") else T["sheetBlurbs"]["decisions"]))
        ws.cell(row=rr, column=4, value=n)
        if n < 1:
            # Header-only sheet (a lane can have an empty section — see the nrows guard in
            # finish()). "$G$2:$G$1" is a reversed range; Excel silently normalises it and
            # counts the header, and stricter readers flag the file for repair. Emit zeros.
            for col in (5, 6, 7):
                ws.cell(row=rr, column=col, value=0)
            continue
        ws.cell(row=rr, column=5, value=f'=COUNTIF({q},"{ok}")')
        ws.cell(row=rr, column=6, value=f'=COUNTIF({q},"{CHW}")')
        ws.cell(row=rr, column=7, value=f'=COUNTIF({q},"{DQW}")')
        ws.cell(row=rr, column=8, value=f"=D{rr}-E{rr}-F{rr}-G{rr}")
        for ci in range(1, 9):
            c = ws.cell(row=rr, column=ci)
            c.font = BODY_B if ci == 2 else BODY; c.border = BORDER
            c.alignment = WRAP if ci == 3 else Alignment(
                vertical="center", horizontal="left" if ci == 2 else "center")
            c.fill = ALT_FILL if i % 2 else REF_FILL
        ws.row_dimensions[rr].height = 28
    tot = start + 1 + len(built)
    ws.cell(row=tot, column=3, value=T["totalLabel"])
    for ci, col in [(4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H")]:
        ws.cell(row=tot, column=ci, value=f"=SUM({col}{start+1}:{col}{tot-1})")
    for ci in range(1, 9):
        c = ws.cell(row=tot, column=ci)
        c.font = BODY_B; c.border = BORDER; c.fill = PatternFill("solid", fgColor="EDE6F5")
        c.alignment = Alignment(vertical="center", horizontal="right" if ci == 3 else "center")
    n = ws.cell(row=tot + 2, column=2, value=T["summaryFootnote"])
    n.font = Font(name=FONT, size=10, italic=True); n.alignment = WRAP
    ws.merge_cells(start_row=tot + 2, start_column=2, end_row=tot + 3, end_column=8)
    ws.sheet_view.showGridLines = False

    order = [S["readme"], S["summary"]] + [name for _, name, _, _ in built]
    wb._sheets = [wb[x] for x in order]
    wb.active = 0
    for s in wb.worksheets:
        s.sheet_properties.tabColor = "7A4E9B"
    wb[S["readme"]].sheet_properties.tabColor = "3B2A5A"
    wb[built[0][1]].sheet_properties.tabColor = "C2185B"

    xlsx = out_dir / f"{stem}.xlsx"
    wb.save(xlsx)

    # ================================================================ text mirror
    def esc(v):
        return "" if v is None else str(v).replace("|", "\\|").replace("\n", "<br>")
    md = [f"# {D['laneLabel']} — {SC['title']} — review packet mirror (v{version})", "",
          f"_Generated by `docs/language-review/pipeline/build_workbook.py` from `{data_path.name}`. "
          f"Text twin of `{stem}.xlsx` — **do not hand-edit**; re-run the pipeline instead. Exists so git "
          f"can diff a packet and so a string can be found without opening Excel._", "",
          f"- Lane `{lane}` · scope `{scope}` · source language `{D['lang']}` · repo `v{version}`",
          f"- {SC['subtitle']}",
          "- Rows: " + " · ".join(f"{k} {len(v)}" for k, v in sect.items() if v), ""]
    if D.get("generatedFrom"):
        md += ["Built from:", ""] + [f"- `{k}` — {v}" for k, v in D["generatedFrom"].items()] + [""]
    if D.get("warnings"):
        md += ["## Extractor warnings", ""] + [f"- {w}" for w in D["warnings"]] + [""]
    for name, headers, rows in mirror:
        md += [f"## {name}", "", "| " + " | ".join(headers) + " |",
               "|" + "|".join(["---"] * len(headers)) + "|"]
        md += ["| " + " | ".join(esc(c) for c in row) + " |" for row in rows]
        md += [""]
    (out_dir / f"{stem}.md").write_text("\r\n".join(md), encoding="utf-8")

    # Committed fingerprint of the repo files this packet was built from. `dev` keeps moving
    # inside a version, so the filename's version number is not enough to tell a fresh packet
    # from a stale one. check_freshness.py reads this.
    (out_dir / f"{stem}.sources.json").write_text(json.dumps({
        "lane": lane, "scope": scope, "version": version,
        # contentHash is the staleness signal: the rows this packet actually contains.
        # builtFrom is provenance only — a source file can change in ways that cannot touch
        # this packet (another language's column landing in the same table).
        "contentHash": D.get("contentHash"),
        # What render_email.mjs reads to state the size of the ask. Recorded here rather than
        # recomputed at send time, so the number in the covering email and the number in the
        # workbook cannot drift apart.
        "rowCounts": {name: n for _, name, n, _ in built},
        # rowTotal counts REVIEW rows and so excludes the decisions sheet, matching what
        # extract.mjs reports and what STATUS.md quotes. A decision is a policy question, not
        # a row of content, and folding the two together would inflate the size of the ask
        # stated in the covering email.
        "rowTotal": sum(len(v) for v in sect.values()),
        "builtFrom": D.get("sourceFingerprint", {}),
    }, indent=1, ensure_ascii=False) + "\n", encoding="utf-8")

    total = sum(len(v) for v in sect.values())
    print(f"{lane} · {scope} (v{version}) — {total} review rows across {len(built)} sheets")
    for _, name, n, _ in built:
        print(f"    {n:>6}  {name}")
    if D.get("warnings"):
        print(f"  ! {len(D['warnings'])} extractor warning(s) carried into the mirror — read before sending.")
    print(f"  {xlsx}\n  {out_dir / f'{stem}.md'}\n  {out_dir / f'{stem}.sources.json'}")
    print("\nNEXT: recalculate so the summary tallies carry cached values, otherwise they read")
    print("      as empty in anything that isn't Excel:")
    print(f"      soffice --headless --convert-to xlsx --outdir {out_dir} {xlsx}")


def _repo_version():
    m = (REPO / "lib" / "version.js").read_text(encoding="utf-8")
    import re
    g = re.search(r'CURRENT_VERSION\s*=\s*"([^"]+)"', m)
    return g.group(1) if g else "0.0.0"


if __name__ == "__main__":
    main()
