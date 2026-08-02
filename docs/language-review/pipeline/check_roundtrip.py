#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_roundtrip.py — prove a reviewer's answers survive the trip out and back.

    python docs/language-review/pipeline/check_roundtrip.py
    python docs/language-review/pipeline/check_roundtrip.py --lane es-latam

For every packet currently sitting in <lane>/template/, this fills the workbook the way a
reviewer would — a verdict, a correction and a note on every data sheet — runs ingest.py over
the result, and checks that every planted answer comes back out in the changeset. Exits
non-zero if anything went missing. Nothing is written outside a temporary directory; the
committed packets are read, never modified.

WHY THIS EXISTS ALONGSIDE check_example.py
`check_example.py` ingests a *frozen fixture* — a workbook committed months ago. That proves
ingest.py still reads the packet shape it was written against. It cannot notice that the
packets you are about to SEND have a different shape, because it never opens them. So a column
added or reordered in build_workbook.py passes check_example.py cleanly and still loses every
correction a reviewer writes into the column after it.

That failure is silent at every step. The reviewer fills the sheet, the file returns, ingest.py
reports a plausible-looking changeset with the missing rows simply absent, and the corrections
are gone with nobody having seen an error. Reviewer attention is the scarce resource in this
whole lane (handoff strategy §7); losing a pass of it to a column index is the most expensive
cheap mistake available here.

The two checks are complements: check_example.py pins the *rendering* (has the changeset's
wording or grouping changed?), this pins the *plumbing* (does today's packet still round-trip?).

WHY THIS IS PYTHON WHEN THE SEND-PATH TOOLS ARE NODE
Because it has to write verdicts into an .xlsx and then run ingest.py, which is itself Python +
openpyxl. A Node version could not do either without shelling out to the Python it was supposed
to avoid needing. The Node/Python split in this pipeline tracks WHO runs a thing, not taste:
send-day tools (check_freshness.mjs, render_email.mjs) are Node because they run on the
maintainer's machine, which has no Python. This is a pipeline-maintenance check that runs
wherever packets are built, so it lives with the rest of the workbook code.

RUN IT after changing build_workbook.py, ingest.py or any sheet layout, and after regenerating
packets — before sending them.

HOW IT FINDS THINGS, so it stays honest
The verdict column is located by its data-validation dropdown and the reviewer's input cells by
their fill colour — both structural properties of the generated workbook. It deliberately does
not reuse ingest.py's own column mapping: a check that shares the assumption it is testing
cannot fail when that assumption is wrong.
"""
import argparse, json, re, shutil, subprocess, sys, tempfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("check_roundtrip.py needs openpyxl:  pip install openpyxl")

HERE = Path(__file__).resolve().parent
REVIEW_ROOT = HERE.parent
# build_workbook.py IN_FILL — the cells a reviewer writes in. Compared on the last six hex
# digits because openpyxl writes the alpha channel as 00 and the LibreOffice recalculation pass
# rewrites it as FF; matching the full eight silently matched nothing on un-recalculated
# packets, and this check then planted nothing and passed.
IN_FILL = "FFF7CC"
COL = re.compile(r"([A-Z]+)\d")


def input_columns(ws):
    """Column indices the reviewer is meant to type into, by fill colour."""
    cols = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=c)
        rgb = getattr(cell.fill.start_color, "rgb", None)
        if rgb and str(rgb)[-6:].upper() == IN_FILL:
            cols.append(c)
    return cols


def verdict_column(ws):
    """The column carrying the verdict dropdown, or None on a non-data sheet."""
    for dv in ws.data_validations.dataValidation:
        for rng in str(dv.sqref).split():
            m = COL.match(rng)
            if m:
                return openpyxl.utils.column_index_from_string(m.group(1))
    return None


def check(lane, xlsx, sources, verbose):
    T = json.loads((HERE / "i18n" / f"{lane}.json").read_text(encoding="utf-8"))
    if "scopes" not in T:
        return None                                   # email-only lane file; nothing to test
    scope, version = sources["scope"], sources["version"]
    S = dict(T["sheets"]); S.update(T["scopes"][scope].get("sheetNames", {}))
    decision_sheets = {S.get("decisions"), S.get("decisionsContent")} - {None}
    # The second verdict word is the dissenting one in both vocabularies. The FIRST
    # ("OK" / "APRUEBO") is a no-op ingest.py skips by design, so planting it would look
    # exactly like data loss.
    change, change_dec = T["verdicts"][1], T["decisionVerdicts"][1]

    tmp = Path(tempfile.mkdtemp())
    sub = tmp / f"2026-01-01-roundtrip-{scope}-v{version}.xlsx"
    shutil.copy(xlsx, sub)
    wb = openpyxl.load_workbook(sub)

    planted, untestable = {}, []
    # The summary sheet carries a sign-off dropdown but no reviewer input cells, and the
    # instructions sheet carries neither. Skipping them by name keeps "dropdown but nothing
    # to plant" meaningful as a warning instead of firing once per packet.
    skip = {S["readme"], S["summary"]}
    for ws in wb.worksheets:
        if ws.title in skip:
            continue
        vc = verdict_column(ws)
        if vc is None or ws.max_row < 2:
            continue
        ins = [c for c in input_columns(ws) if c != vc]
        if not ins:
            # A data sheet with a verdict dropdown but no writable cells is either a broken
            # packet or a broken assumption in this checker. Either way it is a failure: the
            # one thing this script must never do is report success on a sheet it never tested.
            print(f"FAIL  {lane}/{scope}: {ws.title} has a verdict dropdown but no input "
                  f"cells — nothing could be planted, so nothing was verified")
            untestable.append(ws.title)
            continue
        tag = f"RT-{scope}-{ws.title}"
        ws.cell(2, vc, change_dec if ws.title in decision_sheets else change)
        ws.cell(2, ins[0], f"{tag}-FIX")              # first input col = the correction
        ws.cell(2, ins[-1], f"{tag}-NOTE")            # last input col = the note
        planted[ws.title] = tag
    if untestable or not planted:
        print(f"FAIL  {lane}/{scope}: {len(planted)} of {len(planted) + len(untestable)} "
              f"data sheets were testable — refusing to report a pass")
        return False
    wb.save(sub)

    out = tmp / "actual.changeset.md"
    r = subprocess.run([sys.executable, str(HERE / "ingest.py"), str(sub),
                        "--lane", lane, "--scope", scope, "--out", str(out)],
                       capture_output=True, text=True)
    if r.returncode != 0:
        print(f"FAIL  {lane}/{scope}: ingest.py exited {r.returncode}\n{r.stderr}")
        return False
    if "missing expected column" in r.stderr:
        print(f"FAIL  {lane}/{scope}: ingest.py skipped a sheet\n{r.stderr}")
        return False

    text = out.read_text(encoding="utf-8")
    lost = []
    for sheet, tag in planted.items():
        for kind in ("FIX", "NOTE"):
            if f"{tag}-{kind}" not in text:
                lost.append(f"{sheet} ({kind.lower()})")
    if lost:
        print(f"FAIL  {lane}/{scope}: {len(lost)} planted answer(s) did not survive ingest")
        for x in lost:
            print(f"        lost: {x}")
        print(f"      submission kept for inspection: {sub}")
        return False
    print(f"ok    {lane}/{scope}  {len(planted)} sheets round-tripped"
          + (f"  ({', '.join(sorted(planted))})" if verbose else ""))
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--lane", help="only this lane (default: every lane with built packets)")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    found, failures = 0, 0
    for src in sorted(REVIEW_ROOT.glob("*/template/*.sources.json")):
        sources = json.loads(src.read_text(encoding="utf-8"))
        lane = sources["lane"]
        if args.lane and lane != args.lane:
            continue
        xlsx = src.with_name(src.name.replace(".sources.json", ".xlsx"))
        if not xlsx.exists():
            print(f"FAIL  {lane}/{sources['scope']}: manifest with no workbook beside it")
            failures += 1
            continue
        res = check(lane, xlsx, sources, args.verbose)
        if res is None:
            continue
        found += 1
        failures += (not res)

    if not found:
        sys.exit("no built packets found — run extract.mjs + build_workbook.py first")
    print(f"\n{found - failures} round-tripped · {failures} lost data")
    sys.exit(1 if failures else 0)


if __name__ == "__main__":
    main()
