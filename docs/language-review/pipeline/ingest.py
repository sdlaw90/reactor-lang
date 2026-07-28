#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ingest.py — read a returned review workbook and emit a text changeset.

    python docs/language-review/pipeline/ingest.py \
        docs/language-review/es-latam/submitted/2026-08-04-maria-v3.2.0.xlsx

Writes docs/language-review/<lane>/changesets/<same-stem>.md containing every row the
reviewer did not mark OK, grouped by the source file that has to change, plus the open
questions and the decision answers.

THE SUBMITTED FILE IS NEVER MODIFIED OR MOVED. It is the reviewer's testimony and stays
verbatim in submitted/ forever. Status lives in STATUS.md and in implemented/, not in
where the file sits — a submission is almost never wholly applied (some rows land, some
stay open as questions, some are deferred), so a folder can't honestly represent it.

Column lookup is by header text, read from i18n/<lane>.json, so a reordered sheet does
not silently shift the parse.
"""
import argparse, json, re, sys
from collections import defaultdict, Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent.parent


def col_index(ws, header):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=c).value or "").strip() == header:
            return c
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("submission", help="path to the returned .xlsx")
    ap.add_argument("--lane", default=None, help="default: inferred from the path")
    ap.add_argument("--scope", default=None, choices=["interface", "taught", "explanation"],
                    help="default: inferred from the filename")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    sub = Path(args.submission).resolve()
    if not sub.exists():
        sys.exit(f"no such file: {sub}")
    # A lane folder is one carrying STATUS.md. Matching on "is a directory under
    # language-review/" would also match pipeline/, which is not a lane.
    lane = args.lane or next(
        (p.name for p in sub.parents
         if (HERE.parent / p.name / "STATUS.md").exists()), None)
    if not lane:
        sys.exit("could not infer the lane from the path — pass --lane. A lane is a folder "
                 "under docs/language-review/ carrying a STATUS.md.")

    T = json.loads((HERE / "i18n" / f"{lane}.json").read_text(encoding="utf-8"))
    scope = args.scope or next((sc for sc in ("interface", "taught", "explanation")
                                if f"-{sc}-" in sub.name), None)
    OK, CHANGE, QUESTION = T["verdicts"]
    APPROVE = T["decisionVerdicts"][0]
    if not scope:
        sys.exit("could not infer the scope from the filename — pass --scope. A submission's "
                 "name should keep the packet's scope so a return is never ambiguous.")
    SC = T["scopes"][scope]
    S = dict(T["sheets"]); S.update(SC.get("sheetNames", {}))
    H = T["headers"]
    HT, HV, HD = H["text"], H["variants"], H["decisions"]

    wb = load_workbook(sub, data_only=True)

    # ---- reviewer metadata off the summary sheet
    meta = {}
    if S["summary"] in wb.sheetnames:
        ws = wb[S["summary"]]
        for r in range(1, 12):
            label = ws.cell(row=r, column=2).value
            if label and str(label).strip() in T["signoffFields"]:
                meta[str(label).strip()] = ws.cell(row=r, column=3).value

    changes, questions, decisions, advisory = [], [], [], []
    tally = Counter()

    def grab(sheet_key, headers, verdict_h, id_h, cur_h, fix_hs, note_h,
             ctx_hs, file_h=None, scope_h=None):
        name = S[sheet_key]
        if name not in wb.sheetnames:
            return
        ws = wb[name]
        idx = {h: col_index(ws, h) for h in headers}
        missing = [h for h, i in idx.items() if i is None]
        if missing:
            print(f"  ! {name}: missing expected column(s) {missing} — sheet skipped", file=sys.stderr)
            return
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx[verdict_h]).value
            v = str(v).strip() if v is not None else ""
            if not v:
                tally[(name, "—")] += 1
                continue
            tally[(name, v)] += 1
            if v in (OK, APPROVE):
                continue
            rec = {
                "sheet": name,
                "id": ws.cell(row=r, column=idx[id_h]).value,
                "verdict": v,
                "current": ws.cell(row=r, column=idx[cur_h]).value,
                # A row can carry more than one correction field (a variant row has both a
                # corrected term and a corrected country list). With one field the value speaks
                # for itself; with several, label them, or "banano | CO" is a guess.
                "fix": (lambda f: f[0][1] if len(f) == 1 else
                        " · ".join(f"{h.lower()}: {v}" for h, v in f))(
                    [(h, str(ws.cell(row=r, column=idx[h]).value))
                     for h in fix_hs if ws.cell(row=r, column=idx[h]).value]) or "",
                "note": ws.cell(row=r, column=idx[note_h]).value,
                "ctx": {h: ws.cell(row=r, column=idx[h]).value for h in ctx_hs},
                "file": ws.cell(row=r, column=idx[file_h]).value if file_h else None,
                "scope": ws.cell(row=r, column=idx[scope_h]).value if scope_h else None,
            }
            if scope_h and rec["scope"] and T["advisory"] in str(rec["scope"]):
                advisory.append(rec)
            elif sheet_key == "decisions":
                decisions.append(rec)
            elif v == QUESTION:
                questions.append(rec)
            else:
                changes.append(rec)

    if scope == "interface":
        for key in ("variantConfig", "security", "ui", "components", "prose", "changelog"):
            grab(key, HT, HT[5], HT[0], HT[4], [HT[6]], HT[7], [HT[1], HT[2]], file_h=HT[10])
        grab("variants", HV, HV[8], HV[0], HV[4], [HV[9], HV[10]], HV[11],
             [HV[1], HV[2], HV[3], HV[5]], scope_h=HV[7])
        grab("decisions", HD, HD[4], HD[0], HD[2], [HD[5]], HD[6], [HD[1]])
    else:
        HQ, HF, HO = H["questions"], H["fono"], H["overlay"]
        HB = H["vocabTaught"] if scope == "taught" else H["vocabGlossary"]
        grab("questions", HQ, HQ[11], HQ[0], HQ[5], [HQ[12]], HQ[13],
             [HQ[1], HQ[2], HQ[3], HQ[6], HQ[7]], file_h=HQ[14])
        grab("fono", HF, HF[10], HF[0], HF[4], [HF[11]], HF[12],
             [HF[1], HF[3], HF[6], HF[7]], file_h=HF[13])
        if scope == "taught":
            grab("vocab", HB, HB[7], HB[0], HB[2], [HB[8]], HB[9], [HB[3], HB[4], HB[5]], file_h=HB[10])
        else:
            grab("vocab", HB, HB[6], HB[0], HB[4], [HB[7]], HB[8], [HB[2], HB[3], HB[5]], file_h=HB[9])
            grab("overlay", HO, HO[6], HO[0], HO[3], [HO[7]], HO[8], [HO[1], HO[2]], file_h=HO[9])
        grab("decisionsContent", HD, HD[4], HD[0], HD[2], [HD[5]], HD[6], [HD[1]])

    # ---------------------------------------------------------------- write
    out = Path(args.out) if args.out else REPO / "docs" / "language-review" / lane / "changesets" / f"{sub.stem}.md"
    out.parent.mkdir(parents=True, exist_ok=True)

    def block(recs, title, blurb):
        if not recs:
            return [f"## {title}", "", "_None._", ""]
        L = [f"## {title} ({len(recs)})", "", blurb, ""]
        by_file = defaultdict(list)
        for rec in recs:
            by_file[rec["file"] or rec["sheet"]].append(rec)
        for f in sorted(by_file):
            L += [f"### `{f}`", ""]
            for rec in by_file[f]:
                ctx = " · ".join(f"{k}: {v}" for k, v in rec["ctx"].items() if v)
                L += [f"- **{rec['id']}** — {ctx}",
                      f"  - now: `{rec['current']}`"]
                if rec["fix"]:
                    L.append(f"  - proposed: `{rec['fix']}`")
                if rec["note"]:
                    L.append(f"  - note: {rec['note']}")
            L.append("")
        return L

    md = [f"# {lane} · {scope} — changeset from `{sub.name}`", "",
          f"_Generated {date.today().isoformat()} by `docs/language-review/pipeline/ingest.py`. "
          f"Derived from the submission; the submission itself is untouched. Regenerate rather "
          f"than hand-edit._", ""]
    for k, v in meta.items():
        md.append(f"- **{k}** {v if v is not None else '—'}")
    md += ["", "## Verdict tally", "",
           "| Sheet | " + " | ".join([OK, CHANGE, QUESTION, APPROVE, "blank"]) + " |",
           "|---|---|---|---|---|---|"]
    for name in [S[k] for k in (("variants", "decisions", "variantConfig", "security",
                                 "ui", "components", "prose", "changelog")
                                if scope == "interface" else
                                ("questions", "fono", "vocab", "overlay", "decisionsContent"))
                 if k in S and S[k] in wb.sheetnames]:
        md.append("| " + name + " | " + " | ".join(
            str(tally[(name, w)]) for w in [OK, CHANGE, QUESTION, APPROVE, "—"]) + " |")
    md += [""]
    md += block(decisions, "Decisions to apply first",
                "_These are systematic: one answer here can rewrite many rows below. "
                "Apply them before touching individual strings._")
    md += block(changes, "Corrections to apply", "_Reviewer marked these for change._")
    md += block(questions, "Open questions",
                "_Reviewer was unsure or said it depends on the country. Resolve with them; "
                "do not guess._")
    md += block(advisory, "Advisory — outside this lane's scope",
                "_Rows belonging to the counterpart variety. Useful signal, not a sign-off. "
                "Carry them into that lane's packet rather than applying them here._")
    out.write_text("\r\n".join(md), encoding="utf-8")

    print(f"lane {lane} · scope {scope} — {sub.name}")
    print(f"  {len(decisions)} decision answers · {len(changes)} corrections · "
          f"{len(questions)} open questions · {len(advisory)} advisory")
    print(f"  wrote {out}")
    print(f"\nNEXT: apply the decisions first, then the corrections. Record what actually landed in")
    print(f"      docs/language-review/{lane}/implemented/{sub.stem}-applied.md and update STATUS.md.")


if __name__ == "__main__":
    main()
