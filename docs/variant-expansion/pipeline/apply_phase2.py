# -*- coding: utf-8 -*-
"""Apply Phase-2 tense-swap variants: splice into verbo banks, tag, round-trip
JSON + review packet. Mirrors apply_phase1 but for the tense-swap set."""
import json, os, re
import apply_phase1 as a1
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT2 = os.path.join(a1.PIPE, '_out2')
MARK = "// ——— Phase-2 tense-swap variants (depth verbo bank) — PENDING native review ———"
TAGMARK = "// Phase-2 tense-swap variants (depth verbo bank) — PENDING native review"

def splice_track(code, variants):
    path = os.path.join(a1.TRACKS, a1.TRACK[code] + '.js')
    src = open(path, encoding='utf-8').read()
    if MARK in src:
        raise SystemExit(f"[{code}] phase-2 marker already present")
    m = re.search(r'const BANK', src)
    vstart = src.index('verbo: [', m.end())
    vend = src.index('\n  ],', vstart)
    lines = ['    ' + MARK] + [a1.variant_line(v) for v in variants]
    src = src[:vend] + '\n' + '\n'.join(lines) + src[vend:]
    open(path, 'w', encoding='utf-8').write(src)

def add_tags(code, variants):
    path = os.path.join(a1.TRACKS, a1.TRACK[code] + 'Tags.js')
    src = open(path, encoding='utf-8').read()
    if TAGMARK in src:
        raise SystemExit(f"[{code}] phase-2 tag marker already present")
    bt = a1.base_themes_map(code); vp = a1.vid_prompt_map(code)
    rstart = src.index('const RAW = {'); rend = src.index('\n};', rstart)
    lines = ['  ' + TAGMARK]
    for v in variants:
        themes = bt.get(a1.norm(vp.get(v['base_vid'], '')), [])
        parts = []
        if themes:
            parts.append('themes: ' + a1.js(themes))
        parts.append('grammar: T.' + v['tense'])
        parts.append('person: P.' + v['person'])
        lines.append(f'  {a1.js(v["prompt"])}: {{ {", ".join(parts)} }},')
    src = src[:rend] + '\n' + '\n'.join(lines) + src[rend:]
    open(path, 'w', encoding='utf-8').write(src)

def write_roundtrip(code, meta, variants):
    obj = {
        'language': a1.LANGNAME[code], 'track': a1.TRACK[code], 'generated_date': a1.DATE,
        'phase': 'Phase 2 — tense-swap of flexible present-tense verbo items (BULK)',
        'method': ('Recast pronoun-subject present-tense depth items into other tenses '
                   '(past-compound / imperfect / future) with a leading time-marker; person held '
                   'fixed. Distractors = same verb+person across tenses. verbecc'
                   + (' generated, gate OFF (pt)' if code in ('ptBr', 'ptPt') else '-verified (fr/it)')
                   + '. Compound-past skipped for être/essere-auxiliary verbs (participle agreement '
                     'verbecc does not resolve). Excludes already-time-marked, reflexive, '
                     'impersonal, implied/ambiguous-subject items. BULK run overrides the '
                     'methodology hold (sample-only until Phase-1 review); ships to beta pre-review, '
                     'flagged for native review — tense-swaps can read oddly for stative sentences.'),
        'id_scheme': '<code>-verb-<NNN>-t-<tenseKey>',
        'verify_verbecc': meta['verify'],
        'counts': {'present_bases': meta['present_bases'], 'variants': len(variants),
                   'skipped': meta['skips']},
        'variants': [{'vid': v['vid'], 'base_vid': v['base_vid'], 'prompt': v['prompt'],
                      'options': v['options'], 'correctIndex': v['ci'], 'answer': v['options'][v['ci']],
                      'lemma': v['lemma'], 'tense': v['tense'], 'person': v['person'],
                      'level': v['level'], 'explanation': {'en': v['en'], 'es': v['es']},
                      'subtitle_en': v['sub']} for v in variants],
    }
    json.dump(obj, open(os.path.join(a1.GEN, f'{code}_depth_tenseswap.json'), 'w'), ensure_ascii=False, indent=1)

def write_packet(code, variants):
    YEL = PatternFill('solid', fgColor='FFF6CC'); HDR = PatternFill('solid', fgColor='D9D9D9')
    GREY = PatternFill('solid', fgColor='ECECEC')
    TH = Side(style='thin', color='C9C9C9'); BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
    AR = Font(name='Arial', size=10); ARB = Font(name='Arial', size=10, bold=True)
    def cell(ws, r, c, v, fill=None, bold=False):
        x = ws.cell(row=r, column=c, value=v); x.font = ARB if bold else AR
        if fill: x.fill = fill
        x.alignment = Alignment(wrap_text=True, vertical='top'); x.border = BORD; return x
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'READ ME'
    ws.sheet_view.showGridLines = False; ws.column_dimensions['A'].width = 105
    intro = [
        (f"SquirreLingo — {a1.LANGNAME[code]} verb TENSE-swap variants (Phase 2)", True), ("", False),
        ("These take present-tense verb questions and recast them into other tenses (past, "
         "imperfect, future) with a time-word cue (Yesterday / In the past / Tomorrow). Already live in beta.", False),
        ("Fill the yellow columns: Correct? (Y/N) · Correction · Natural? (Y/N) · Notes. Ignore grey ID.", True),
        ("PLEASE FLAG any that sound unnatural — a present sentence pushed into the past/future can "
         "read oddly (e.g. an unchanging state). Those are exactly what this review is for.", True),
    ]
    r = 1
    for t, b in intro:
        cell(ws, r, 1, t, bold=bool(b)); r += 1
    cols = [('ID', 16), ('Question', 44), ('Answer', 20), ('All options', 34), ('Tense', 16),
            ('Person', 12), ('CEFR', 6), ('Correct?', 10), ('Correction', 16), ('Natural?', 10), ('Notes', 20)]
    s = wb.create_sheet('Tense variants'); s.sheet_view.showGridLines = False
    for i, (c, w) in enumerate(cols, 1):
        cell(s, 1, i, c, fill=HDR, bold=True); s.column_dimensions[get_column_letter(i)].width = w
    rr = 2
    for v in sorted(variants, key=lambda x: (x['base_vid'], x['tense'])):
        vals = [v['vid'], v['prompt'], v['options'][v['ci']], ' / '.join(v['options']),
                v['tense'], v['person'], v['level']]
        for i, val in enumerate(vals, 1):
            cell(s, rr, i, val, fill=GREY)
        for i in range(len(vals) + 1, len(vals) + 5):
            cell(s, rr, i, '', fill=YEL)
        rr += 1
    s.freeze_panes = 'A2'; s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{rr-1}"
    wb.save(os.path.join(a1.PACK, f'{code}-depth-tenseswap.xlsx'))

def main():
    summary = {}
    for code in ['fr', 'frCa', 'it', 'ptBr', 'ptPt']:
        meta = json.load(open(os.path.join(OUT2, f'{code}_phase2_variants.json')))
        V = meta['variants']
        splice_track(code, V); add_tags(code, V); write_roundtrip(code, meta, V); write_packet(code, V)
        bt = {}
        for v in V:
            bt[v['tense']] = bt.get(v['tense'], 0) + 1
        summary[code] = {'variants': len(V), 'present_bases': meta['present_bases'], 'by_tense': bt}
        print(f"[{code}] spliced {len(V)} tense variants | bases {meta['present_bases']} | {bt}")
    total = sum(s['variants'] for s in summary.values())
    json.dump(summary, open(os.path.join(OUT2, 'apply2_summary.json'), 'w'), ensure_ascii=False, indent=1)
    print("TOTAL phase-2 variants:", total)

if __name__ == '__main__':
    main()
