# -*- coding: utf-8 -*-
"""
Apply Phase-1 person-swap variants to the tracks:
  - splice variants into <track>.js verbo bank (with // <vid> comments)
  - add tag entries to <track>Tags.js RAW (themes inherited from base + grammar + person)
  - write round-trip JSON  generated/<code>_depth_variants.json
  - write review packet     review-packets/<code>-depth-variants.xlsx
Operates on the staged repo copy in-place. Idempotent guard: refuses if the
variant marker block already present.
"""
import json, re, os, sys, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PIPE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(PIPE, "..", "..", ".."))
TRACKS = os.path.join(REPO, "data", "tracks")
GEN = os.path.join(PIPE, "..", "generated")
PACK = os.path.join(PIPE, "..", "review-packets")
OUT = os.path.join(PIPE, "_out")

TRACK = {'fr': 'frForEn', 'frCa': 'frCaForEn', 'it': 'itForEn', 'ptBr': 'ptBrForEn', 'ptPt': 'ptPtForEn'}
LANGNAME = {'fr': 'French (France)', 'frCa': 'French (Québec)', 'it': 'Italian',
            'ptBr': 'Portuguese (Brazil)', 'ptPt': 'Portuguese (Portugal)'}
DATE = '2026-07-20'

def norm(s):
    return unicodedata.normalize('NFC', re.sub(r'\s+', ' ', str(s)).strip())

def js(v):
    if v is None: return 'null'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, str): return json.dumps(v, ensure_ascii=False)
    if isinstance(v, list): return '[' + ', '.join(js(x) for x in v) + ']'
    if isinstance(v, dict): return '{ ' + ', '.join(f'{k}: {js(val)}' for k, val in v.items()) + ' }'
    raise TypeError(str(type(v)))

def variant_line(v):
    return ('    [' + js(v['prompt']) + ', ' + js(v['options']) + ', ' + str(v['ci']) + ', '
            + js({'en': v['en'], 'es': v['es']}) + ', ' + js(v['level']) + ', null, '
            + js({'en': v['sub']}) + '],  // ' + v['vid'])

def base_themes_map(code):
    """base prompt(norm) -> list[themes] from the existing Tags.js RAW."""
    src = open(os.path.join(TRACKS, TRACK[code] + 'Tags.js'), encoding='utf-8').read()
    out = {}
    for m in re.finditer(r'"((?:[^"\\]|\\.)*)":\s*\{([^}]*)\}', src):
        prompt = m.group(1).replace('\\"', '"').replace('\\\\', '\\')
        th = re.search(r'themes:\s*(\[[^\]]*\])', m.group(2))
        if th:
            try:
                out[norm(prompt)] = json.loads(th.group(1))
            except Exception:
                pass
    return out

def vid_prompt_map(code):
    d = json.load(open(os.path.join(GEN, f'{code}_depth_generated.json')))
    return {it['vid']: it['prompt'] for it in d['verbo']}

MARK = "// ——— Phase-1 person-swap variants (depth verbo bank) — PENDING native review ———"

def splice_track(code, variants):
    path = os.path.join(TRACKS, TRACK[code] + '.js')
    src = open(path, encoding='utf-8').read()
    if MARK in src:
        raise SystemExit(f"[{code}] variant marker already present — refusing to double-splice")
    m = re.search(r'const BANK', src)
    vstart = src.index('verbo: [', m.end())
    vend = src.index('\n  ],', vstart)
    lines = [ '    ' + MARK ] + [variant_line(v) for v in variants]
    block = '\n' + '\n'.join(lines)
    newsrc = src[:vend] + block + src[vend:]
    open(path, 'w', encoding='utf-8').write(newsrc)
    return newsrc.count('\n')  # not meaningful; just touch

def add_tags(code, variants):
    path = os.path.join(TRACKS, TRACK[code] + 'Tags.js')
    src = open(path, encoding='utf-8').read()
    if 'Phase-1 person-swap' in src:
        raise SystemExit(f"[{code}] tag marker already present — refusing to double-add")
    bt = base_themes_map(code)
    vp = vid_prompt_map(code)
    # find RAW closing brace
    rstart = src.index('const RAW = {')
    rend = src.index('\n};', rstart)
    lines = ['  // Phase-1 person-swap variants (depth verbo bank) — PENDING native review']
    for v in variants:
        base_prompt = vp.get(v['base_vid'], '')
        themes = bt.get(norm(base_prompt), [])
        parts = []
        if themes:
            parts.append('themes: ' + js(themes))
        parts.append('grammar: T.' + v['tense'])
        parts.append('person: P.' + v['person'])
        lines.append(f'  {js(v["prompt"])}: {{ {", ".join(parts)} }},')
    block = '\n' + '\n'.join(lines)
    newsrc = src[:rend] + block + src[rend:]
    open(path, 'w', encoding='utf-8').write(newsrc)

def write_roundtrip(code, meta, variants):
    os.makedirs(GEN, exist_ok=True)
    obj = {
        'language': LANGNAME[code], 'track': TRACK[code], 'generated_date': DATE,
        'phase': 'Phase 1 — person-swap of the NEW depth verbo items',
        'method': ('Mechanical person-swap of pronoun-subject standard verbo items from the '
                   'content-depth pass. Conjugations via verbecc'
                   + (' (generated; verbecc gate OFF for Portuguese — forms rely on generate + native review)'
                      if code in ('ptBr', 'ptPt')
                      else ' (verbecc-verified against the native-checked base answer)')
                   + '. Options = sibling person-forms in the same tense; deduped vs the whole bank. '
                     'Reflexive / progressive / impersonal / noun-subject / implied-subject items and the '
                     '1st-person-singular cell of subjunctive items (coreference) were excluded.'),
        'id_scheme': '<code>-verb-<NNN>-p-<person> (base id + -p-<person>)',
        'verify_verbecc': meta['verify'],
        'counts': {'eligible_bases': meta['eligible_bases'], 'variants': len(variants),
                   'skipped': meta['skips']},
        'variants': [{'vid': v['vid'], 'base_vid': v['base_vid'], 'prompt': v['prompt'],
                      'options': v['options'], 'correctIndex': v['ci'],
                      'answer': v['options'][v['ci']], 'lemma': v['lemma'],
                      'tense': v['tense'], 'person': v['person'], 'level': v['level'],
                      'explanation': {'en': v['en'], 'es': v['es']}, 'subtitle_en': v['sub']}
                     for v in variants],
    }
    p = os.path.join(GEN, f'{code}_depth_variants.json')
    json.dump(obj, open(p, 'w'), ensure_ascii=False, indent=1)

def write_packet(code, variants):
    os.makedirs(PACK, exist_ok=True)
    YEL = PatternFill('solid', fgColor='FFF6CC'); HDR = PatternFill('solid', fgColor='D9D9D9')
    GREY = PatternFill('solid', fgColor='ECECEC')
    TH = Side(style='thin', color='C9C9C9'); BORD = Border(left=TH, right=TH, top=TH, bottom=TH)
    AR = Font(name='Arial', size=10); ARB = Font(name='Arial', size=10, bold=True)
    def cell(ws, r, c, v, fill=None, bold=False):
        x = ws.cell(row=r, column=c, value=v); x.font = ARB if bold else AR
        if fill: x.fill = fill
        x.alignment = Alignment(wrap_text=True, vertical='top'); x.border = BORD; return x
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'READ ME'
    ws.sheet_view.showGridLines = False; ws.column_dimensions['A'].width = 105
    intro = [
        (f"SquirreLingo — {LANGNAME[code]} verb person-swap variants (Phase 1)", True), ("", False),
        ("These are existing depth verb questions re-cast into the other subject persons "
         "(I / you / he·she / we / they), same sentence and tense. Already live in beta.", False),
        ("For each row, fill the yellow columns: Correct? (Y/N) · Correction · Natural? (Y/N) · Notes. "
         "Ignore the grey ID column.", True),
        ("Portuguese note: conjugations were machine-generated without the usual verb-checker gate "
         "(it is unreliable for Portuguese) — please check the forms with extra care.",
         code in ('ptBr', 'ptPt')),
    ]
    r = 1
    for t, b in intro:
        if t == "" or t:
            cell(ws, r, 1, t, bold=bool(b)); r += 1
    cols = [('ID', 15), ('Question', 42), ('Answer', 20), ('All options', 34), ('Tense', 16),
            ('Person', 14), ('CEFR', 6), ('Correct?', 10), ('Correction', 16), ('Natural?', 10), ('Notes', 20)]
    s = wb.create_sheet('Person variants'); s.sheet_view.showGridLines = False
    for i, (c, w) in enumerate(cols, 1):
        cell(s, 1, i, c, fill=HDR, bold=True); s.column_dimensions[get_column_letter(i)].width = w
    rr = 2
    for v in sorted(variants, key=lambda x: (x['base_vid'], x['person'])):
        vals = [v['vid'], v['prompt'], v['options'][v['ci']], ' / '.join(v['options']),
                v['tense'], v['person'], v['level']]
        for i, val in enumerate(vals, 1):
            cell(s, rr, i, val, fill=GREY)
        for i in range(len(vals) + 1, len(vals) + 5):
            cell(s, rr, i, '', fill=YEL)
        rr += 1
    s.freeze_panes = 'A2'; s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{rr-1}"
    wb.save(os.path.join(PACK, f'{code}-depth-variants.xlsx'))

def main():
    summary = {}
    for code in ['fr', 'frCa', 'it', 'ptBr', 'ptPt']:
        meta = json.load(open(os.path.join(OUT, f'{code}_phase1_variants.json')))
        variants = meta['variants']
        splice_track(code, variants)
        add_tags(code, variants)
        write_roundtrip(code, meta, variants)
        write_packet(code, variants)
        # per-person breakdown
        bp = {}
        for v in variants:
            bp[v['person']] = bp.get(v['person'], 0) + 1
        summary[code] = {'variants': len(variants), 'eligible_bases': meta['eligible_bases'],
                         'by_person': bp}
        print(f"[{code}] spliced {len(variants)} variants | bases {meta['eligible_bases']} | {bp}")
    json.dump(summary, open(os.path.join(OUT, 'apply_summary.json'), 'w'), ensure_ascii=False, indent=1)
    total = sum(s['variants'] for s in summary.values())
    print("TOTAL variants:", total)

if __name__ == '__main__':
    main()
