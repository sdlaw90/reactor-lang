# -*- coding: utf-8 -*-
import json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
CFG=json.load(open(sys.argv[1])); A=json.load(open(CFG['assembled']))
code=CFG['code']; LANG=CFG['lang_name']; DATE='2026-07-19'
G=f'out_{code}/generated'; R=f'out_{code}/review-packets'; DO=f'out_{code}/docs'
for d in (G,R,DO): os.makedirs(d,exist_ok=True)
obj={'language':LANG,'track':CFG['track'],'generated_date':DATE,
 'method':'Content-depth pass; per-(category×CEFR-band) workflow with native-speaker verify; verbo verbecc-checked; deduped.',
 'counts':{k:len(A[k+'_new']) for k in ('vocab','verbo','trad','fono')} | {'dropped':{k:len(v) for k,v in A['stats']['dropped'].items()}},
 'vocab':A['vocab_new'],'verbo':A['verbo_new'],'trad':A['trad_new'],'fono':A['fono_new'],'tags_count':len(A['tags'])}
json.dump(obj,open(f'{G}/{code}_depth_generated.json','w'),ensure_ascii=False,indent=1)
YEL=PatternFill('solid',fgColor='FFF6CC');HDR=PatternFill('solid',fgColor='D9D9D9');GREY=PatternFill('solid',fgColor='ECECEC')
TH=Side(style='thin',color='C9C9C9');BORD=Border(left=TH,right=TH,top=TH,bottom=TH);AR=Font(name='Arial',size=10);ARB=Font(name='Arial',size=10,bold=True)
def cell(ws,r,c,v,fill=None,bold=False):
    x=ws.cell(row=r,column=c,value=v);x.font=ARB if bold else AR
    if fill:x.fill=fill
    x.alignment=Alignment(wrap_text=True,vertical='top');x.border=BORD;return x
wb=openpyxl.Workbook(); ws=wb.active; ws.title='READ ME'; ws.sheet_view.showGridLines=False; ws.column_dimensions['A'].width=100
for i,(t,b) in enumerate([(f"SquirreLingo — {LANG} content-depth review",1),("",0),
 ("New auto-generated + native-verified practice questions, already live in beta.",0),
 ("For each row fill the yellow columns: Correct? (Y/N) · Correction · Natural? (Y/N) · Notes. Ignore grey ID.",1),
 ("One tab per category. Verbs were also machine-checked with a conjugator; please confirm naturalness.",0)],1):
    cell(ws,i+1,1,t,bold=bool(b))
def sheet(name,rows,cols,getv):
    s=wb.create_sheet(name); s.sheet_view.showGridLines=False
    for i,(c,w) in enumerate(cols,1): cell(s,1,i,c,fill=HDR,bold=True); s.column_dimensions[get_column_letter(i)].width=w
    r=2
    for it in rows:
        vals=getv(it)
        for i,v in enumerate(vals,1): cell(s,r,i,v,fill=GREY)
        for i in range(len(vals)+1,len(vals)+5): cell(s,r,i,'',fill=YEL)
        r+=1
    s.freeze_panes='A2'; s.auto_filter.ref=f"A1:{get_column_letter(len(cols)+4)}{r-1}"
mc=[('ID',13),('Question',40),('Answer',24),('All options',36),('CEFR',6),('Correct?',10),('Correction',16),('Natural?',10),('Notes',20)]
sheet('Vocab',A['vocab_new'],mc,lambda d:[d['vid'],d['prompt'],d['options'][d['ci']],' / '.join(d['options']),d['level']])
sheet('Verbs',A['verbo_new'],mc,lambda d:[d['vid'],d['prompt'],d['options'][d['ci']],' / '.join(d['options']),d['level']])
sheet('Idioms',A['trad_new'],mc,lambda d:[d['vid'],d['prompt'],d['options'][d['ci']],' / '.join(d['options']),d['level']])
fc=[('Sentence',34),('Sound',24),('Diff',6),('Identify ans',30),('Respond ans',26),('Correct?',10),('Correction',16),('Natural?',10),('Notes',20)]
sheet('Phonetics',A['fono_new'],fc,lambda f:[f['text'],f['sound'],f['difficulty'],f['io'][f['ici']],f['ro'][f['rci']]])
wb.save(f'{R}/{code}-depth-review.xlsx')
tot={'vocab':len(A['vocab_new']),'verbo':len(A['verbo_new']),'trad':len(A['trad_new']),'fono':len(A['fono_new'])}
open(f'{DO}/{code}-depth-reference.md','w').write(f"""# {LANG} ({CFG['track']}) — Content-Depth Pass Reference

_Ran {DATE}. Brings {CFG['track']} to the Spanish depth standard; new dedicated
`verbo` ("{CFG['verbo_cat']}") category._

## Result
- Vocab +{tot['vocab']}, Verbs +{tot['verbo']} depth (+{len(A['verbo_migrated'])} migrated variants/base),
  Idioms +{tot['trad']}, Phonetics +{tot['fono']}. Tags: {len(A['tags'])}.
- Dropped in QA: {json.dumps({k:len(v) for k,v in A['stats']['dropped'].items()})} (dups + verbecc conjugation mismatches).

## Method
Per-(category × CEFR band) workflow (~44 agents) + native-speaker adversarial
verify; verbo forms verbecc-verified; deduped vs existing bank; new
`{CFG['tags_file']}` (themes + per-verb tense/person). `gameEngine.js` unchanged
(category-agnostic). AI-authored → beta pre-review; native review gates real-prod.

## Deliverables
- `review-packets/{code}-depth-review.xlsx` — native-review packet.
- `generated/{code}_depth_generated.json` — round-trip data (IDs join to track).
""")
print("deliverables:", tot, "->", f"out_{code}/")
